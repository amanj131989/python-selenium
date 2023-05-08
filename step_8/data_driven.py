# $ pip install openpyxl
# https://docs.google.com/spreadsheets/d/1WfdY4l7RIw5BJEqseWYYsga-txfevDp1fRynjgQqI-U/edit?usp=sharing
import time

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.edge.service import Service as EdgeService
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium import webdriver

from openpyxl import load_workbook


# Login if login page is prompted
def login(driver, u="Admin", p="admin123"):
    if "login" in driver.current_url:
        driver.find_element(By.XPATH, "//input[@name='username']").send_keys(u)
        driver.find_element(By.XPATH, "//input[@name='password']").send_keys(p)
        driver.find_element(By.XPATH, "//button[@type='submit']").click()


# Logout the logged in user
def logout(driver):
    driver.find_element(By.XPATH, "//img[@alt='profile picture']").click()
    driver.find_element(By.XPATH, "//a[@href='/web/index.php/auth/logout']").click()


driver = webdriver.Edge(service=EdgeService(EdgeChromiumDriverManager().install()))

driver.implicitly_wait(4)
driver.maximize_window()

wb_obj = load_workbook(filename="C:/Users/AJiddewar/OneDrive - Quinstreet, Inc/Documents/OtherProjects/TestProj/tests/Scripts/Data-Driven-Example/data_for_data_driven.xlsx")
# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/pim/addEmployee")

for row in sheet_obj.iter_rows():
    if row[0].value and not row[0].value.startswith('#'):
        login(driver)
        driver.find_element(By.XPATH, "//input[@name='firstName']").send_keys(row[0].value)
        driver.find_element(By.XPATH, "//input[@name='middleName']").send_keys(row[1].value)
        driver.find_element(By.XPATH, "//input[@name='lastName']").send_keys(row[2].value)
        # driver.find_element(By.XPATH, "//form//div[contains(@class, 'oxd-grid-2')]//input").send_keys(int(row[3].value))
        if row[4].value:
            driver.find_element(By.XPATH, "//span[contains(@class, 'oxd-switch-input')]").click()
            username, password = row[5].value, row[7].value
            driver.find_element(By.XPATH, "//label[normalize-space()='Username']/ancestor::*[2]//input").send_keys(username)
            driver.find_element(By.XPATH, "//label[normalize-space()='Password']/ancestor::*[2]//input").send_keys(password)
            driver.find_element(By.XPATH, "//label[normalize-space()='Confirm Password']/ancestor::*[2]//input").send_keys(row[7].value)
            driver.find_element(By.XPATH, "//button[@type='submit']").click()
            login(driver, username, password)
            logout(driver)
            # check the login
            driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/pim/addEmployee")

        driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/pim/addEmployee")

time.sleep(10)
driver.quit()
