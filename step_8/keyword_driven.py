# $ pip install openpyxl
# https://docs.google.com/spreadsheets/d/1XQbaXkSJPXgYRqXR1aM96RYUiF8qQCWnzSN3z8WePao/edit?usp=sharing
import time

from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service as EdgeService
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium import webdriver

from openpyxl import load_workbook


wb_obj = load_workbook(filename="C:/Users/AJiddewar/OneDrive - Quinstreet, Inc/Documents/OtherProjects/TestProj/tests/Scripts/Data-Driven-Example/keyword_driven.xlsx")
# Get workbook active sheet object
# from the active attribute
for row in wb_obj.active.iter_rows():
    if row[0].value and not row[0].value.startswith('#'):
        keyword = row[1].value
        if keyword == 'openBrowser':
            driver = webdriver.Edge(service=EdgeService(EdgeChromiumDriverManager().install()))
            driver.implicitly_wait(4)
            driver.maximize_window()
        elif keyword == 'navigate':
            driver.get(row[2].value)
        elif keyword == 'enterUsername':
            driver.find_element(By.XPATH, "//input[@name='username']").send_keys(row[2].value)
        elif keyword == 'enterPassword':
            driver.find_element(By.XPATH, "//input[@name='password']").send_keys(row[2].value)
        elif keyword == 'submit':
            driver.find_element(By.XPATH, "//button[@type='submit']").click()

driver.quit()
